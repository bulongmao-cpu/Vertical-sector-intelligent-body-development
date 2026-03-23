from openpyxl import load_workbook
from collections import defaultdict
import json
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine, text
from sqlalchemy import text
import math
from sklearn.tree import DecisionTreeRegressor
from scipy.stats import f_oneway
import mapclassify
from toolbox import update_ui, CatchException, report_exception, write_history_to_file, promote_file_to_downloadzone
import pandas as pd
import numpy as np
from sklearn.linear_model import LinearRegression
from crazy_functions.crazy_utils import request_gpt_model_in_new_thread_with_ui_alive
import os
import glob
import matplotlib.pyplot as plt
import seaborn as sns
import markdown2
from docx import Document
from docx.shared import Pt, Inches
from docx.oxml.ns import qn
from bs4 import BeautifulSoup
import matplotlib
import re
from sklearn.preprocessing import StandardScaler
from collections import defaultdict, Counter
from matplotlib.font_manager import FontProperties
from crazy_functions.crazy_utils import request_input_from_user

class MySQLWriter:
    def __init__(self, database: str):
        """
        еҲқе§ӢеҢ– MySQLWriter
        """
        self.database = database
        self.engine = create_engine(
            f"mysql+pymysql://root:123456@localhost:3306/{database}?charset=utf8mb4"
        )

    # ==================== еўһпјҲInsertпјү ====================
    def insert_df(self, df: pd.DataFrame, table_name: str, if_exists: str = "replace"):
        df.to_sql(name=table_name, con=self.engine, if_exists=if_exists, index=False)
        print(f"жҸ’е…Ҙ {len(df)} иЎҢж•°жҚ®еҲ° {table_name} пјҲжЁЎејҸпјҡ{if_exists}пјү")

    def insert_one(self, table_name: str, data: dict):
        cols = ", ".join(data.keys())
        placeholders = ", ".join([f":{k}" for k in data.keys()])
        sql = f"INSERT INTO {table_name} ({cols}) VALUES ({placeholders})"
        with self.engine.connect() as conn:
            conn.execute(text(sql), data)
            conn.commit()
        print(f"жҸ’е…ҘеҚ•жқЎж•°жҚ®еҲ° {table_name}")

    # ==================== жҹҘпјҲSelectпјү ====================
    def select_df(self, table_name: str, columns: list[str] = None, where: str = None) -> pd.DataFrame:
        def quote_identifier(name):
            return f"`{name}`"

        if columns is None:
            col_str = "*"
        else:
            col_str = ", ".join(quote_identifier(c) for c in columns)

        table_name = quote_identifier(table_name)   # еӨ„зҗҶеҢ…еҗ«дёӯж–Ү/з©әж ј/зү№ж®Ҡеӯ—з¬Ұзҡ„иЎЁеҗҚ
        sql = f"SELECT {col_str} FROM {table_name}"

        if where:
            sql += f" WHERE {where}"

        df = pd.read_sql(sql, self.engine)
        print(f"д»Һ {table_name} иҜ»еҸ– {len(df)} иЎҢж•°жҚ®")
        return df
    
    # ==================== ж”№пјҲUpdateпјү ====================
    def update(self, table_name: str, updates: dict, where: str):
        set_clause = ", ".join([f"{col} = :{col}" for col in updates.keys()])
        sql = f"UPDATE {table_name} SET {set_clause} WHERE {where}"
        with self.engine.connect() as conn:
            conn.execute(text(sql), updates)
            conn.commit()
        print(f"жӣҙж–° {table_name} иЎЁи®°еҪ•пјҲжқЎд»¶пјҡ{where}пјү")

    # ==================== еҲ пјҲDeleteпјү ====================
    def delete(self, table_name: str, where: str):
        sql = f"DELETE FROM {table_name} WHERE {where}"
        with self.engine.connect() as conn:
            conn.execute(text(sql))
            conn.commit()
        print(f"д»Һ {table_name} еҲ йҷӨи®°еҪ•пјҲжқЎд»¶пјҡ{where}пјү")

    # ==================== еӨҮжіЁ ====================
    def add_table_comment(self, table_name, comment):
        # еҸҚеј•еҸ·еҢ…иЈ№иЎЁеҗҚпјҢйҒҝе…Қдёӯж–Ү/зү№ж®Ҡеӯ—з¬ҰеҮәй”ҷ
        safe_table = f"`{table_name}`"
        # иҪ¬д№үеҚ•еј•еҸ·
        safe_comment = comment.replace("'", "''")
        sql = f"ALTER TABLE {safe_table} COMMENT = '{safe_comment}'"
        with self.engine.begin() as conn:
            conn.execute(text(sql))

    def add_column_comment(self, table_name: str, column_name: str, col_type: str, comment: str):
        with self.engine.connect() as conn:
            sql = f"ALTER TABLE {table_name} MODIFY {column_name} {col_type} COMMENT :comment"
            conn.execute(text(sql), {"comment": comment})
            conn.commit()
        print(f"еҲ— {column_name} е·Іж·»еҠ еӨҮжіЁпјҡ{comment}")

    def get_columns(self, table_name: str) -> list[str]:
        """
        иҝ”еӣһжҢҮе®ҡиЎЁзҡ„жүҖжңүеҲ—еҗҚ
        """
        sql = f"SHOW COLUMNS FROM `{table_name}`"
        df = pd.read_sql(sql, self.engine)
        return df['Field'].tolist()

    # ==================== жҳҫзӨәжүҖжңүиЎЁ ====================
    def show_tables(self) -> list[str]:
        """
        иҺ·еҸ–еҪ“еүҚж•°жҚ®еә“дёӢзҡ„жүҖжңүиЎЁеҗҚ
        """
        sql = "SHOW TABLES"
        df = pd.read_sql(sql, self.engine)
        tables = df.iloc[:, 0].tolist()
        print(f"еҪ“еүҚж•°жҚ®еә“({self.database})е…ұжңү {len(tables)} еј иЎЁ: {tables}")
        return tables

    def find_tables_like(self, keyword: str):
        """
        ж №жҚ®е…ій”®еӯ—жЁЎзіҠжҗңзҙўиЎЁеҗҚпјҢиҝ”еӣһеҢ…еҗ«иҜҘе…ій”®еӯ—зҡ„жүҖжңүиЎЁеҗҚ
        """
        sql = text("""
            SELECT table_name 
            FROM information_schema.tables 
            WHERE table_schema = :db 
              AND table_name LIKE :pattern
        """)
        
        with self.engine.connect() as conn:
            result = conn.execute(sql, {"db": self.database, "pattern": f"%{keyword}%"})
            tables = [row[0] for row in result.fetchall()]
        return tables
    
    # ==================== йҖҡз”Ёж–№жі• ====================
    def execute_sql(self, sql: str, params: dict = None):
        with self.engine.connect() as conn:
            conn.execute(text(sql), params or {})
            conn.commit()
        print(f"жү§иЎҢ SQL: {sql}")


    def aggregate_for_key(self, row_name, year, table_list, require_year_for_tables):
        """
        row_name: иЎҢеҗҚпјҲеҸҜд»Ҙдёә Noneпјү
        year: '2020е№ҙ' жҲ– None
        table_list: ['жҜҚйёЎ_еҹәзЎҖиЎЁ3  дәәе‘ҳж”ҜеҮә', ...]
        require_year_for_tables: list of booleans, дёҺ table_list еҜ№еә”пјҲTrue иЎЁзӨәиҜҘиЎЁеңЁеҲ—зӯӣйҖүж—¶йңҖиҰҒеҢ…еҗ« yearпјү

        иҝ”еӣһ (row_name, year, total_amount)
        """
        if not table_list:
            return (row_name, year, 0)

        total_sum = 0.0
        with self.engine.connect() as conn:
            for table, need_year in zip(table_list, require_year_for_tables):
                # --- иҺ·еҸ–йҮ‘йўқеҲ— ---
                if need_year and year:
                    q = text("""
                        SELECT COLUMN_NAME
                        FROM information_schema.columns
                        WHERE table_name = :table
                        AND (
                            COLUMN_NAME LIKE '%йҮ‘йўқ%' OR
                            COLUMN_NAME LIKE '%иҙ№%' OR
                            COLUMN_NAME LIKE '%е°Ҹи®Ў%' OR
                            COLUMN_NAME LIKE '%еҗҲи®Ў%'
                        )
                        AND COLUMN_NAME LIKE :year_like
                    """)
                    params = {'table': table, 'year_like': f"%{year}%"}
                else:
                    q = text("""
                        SELECT COLUMN_NAME
                        FROM information_schema.columns
                        WHERE table_name = :table
                        AND (
                            COLUMN_NAME LIKE '%йҮ‘йўқ%' OR
                            COLUMN_NAME LIKE '%иҙ№%' OR
                            COLUMN_NAME LIKE '%е°Ҹи®Ў%' OR
                            COLUMN_NAME LIKE '%еҗҲи®Ў%'
                        )
                    """)
                    params = {'table': table}

                cols = [r[0] for r in conn.execute(q, params)]
                if not cols:
                    print(f"иЎЁ {table} жңӘжүҫеҲ°йҮ‘йўқеҲ—пјҢи·іиҝҮгҖӮ")
                    continue

                # --- иҮӘеҠЁиҺ·еҸ–з¬¬дәҢеҲ—еӯ—ж®өеҗҚ ---
                q2 = text("""
                    SELECT COLUMN_NAME
                    FROM information_schema.columns
                    WHERE table_name = :table
                    ORDER BY ORDINAL_POSITION
                    LIMIT 1 OFFSET 1
                """)
                second_col = conn.execute(q2, {'table': table}).scalar()
                if not second_col:
                    print(f"иЎЁ {table} ж— з¬¬дәҢеҲ—пјҢи·іиҝҮгҖӮ")
                    continue

                # --- и®Ўз®—йҮ‘йўқеҗҲи®Ў ---
                sum_expr = " + ".join([f"COALESCE(`{c}`,0)" for c in cols])
                if row_name:
                    sql = text(f"""
                        SELECT SUM({sum_expr}) AS total
                        FROM `{table}`
                        WHERE `{second_col}` LIKE :kw
                    """)

                    # жЁЎзіҠе…ій”®иҜҚжҸҗеҸ–пјҡеҺ»жҺүвҖңйғЁй—ЁвҖқзӯүеӯ—е°ҫеҶҚжЁЎзіҠеҢ№й…Қ
                    kw = row_name.replace('ж”ҜеҮә', '').replace('дәәе‘ҳ', '') if row_name else ''
                    res = conn.execute(sql, {'kw': f"%{kw}%"}).fetchone()
                else:
                    sql = text(f"""
                        SELECT SUM({sum_expr}) AS total
                        FROM `{table}`
                    """)
                    res = conn.execute(sql).fetchone()

                subtotal = float(res[0]) if res and res[0] is not None else 0.0
                if row_name:
                    print(f"иЎЁ {table} | з¬¬дәҢеҲ—={second_col} | е…ій”®иҜҚ='{kw}' | жұҮжҖ»йҮ‘йўқ={subtotal}")
                else:
                    print(f"иЎЁ {table} | з¬¬дәҢеҲ—={second_col} | жұҮжҖ»йҮ‘йўқ={subtotal}")
                total_sum += subtotal

        print(f"жұҮжҖ»е®ҢжҲҗпјҡrow_name={row_name}, year={year}, total={total_sum}")
        return (row_name, year, total_sum)

class SheetMapper:
    def __init__(self, df: pd.DataFrame):
        """
        еҲқе§ӢеҢ–пјҡиҫ“е…ҘеҺҹе§Ӣ DataFrameпјҲеёҰжңү 'ж”ҜеҮәзұ»еһӢ' еӨҡеҲ—пјү
        """
        # еҸ–з¬¬дёҖиЎҢдҪңдёәиЎЁеӨҙ
        header = df.iloc[0]
        self.df = df[1:].reset_index(drop=True)
        self.df.columns = header

        # жүҫеҮәвҖңж”ҜеҮәзұ»еһӢвҖқзӣёе…іеҲ—
        self.row_name_cols = [col for col in self.df.columns if "ж”ҜеҮәзұ»еһӢ" in str(col)]

        # ж·»еҠ е”ҜдёҖиЎҢеҗҚеҲ—
        self.df["иЎҢеҗҚ"] = self.df[self.row_name_cols].apply(self._make_row_name, axis=1)

        # еҲқе§ӢеҢ–еӯ—е…ё
        self.data_dict = self._to_dict()

    def _make_row_name(self, row):
        """
        з”ҹжҲҗе”ҜдёҖиЎҢеҗҚпјҡеҺ»жҺүйҮҚеӨҚе’Ңз©әеҖј
        """
        values = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() != ""]
        values = list(dict.fromkeys(values))  # еҺ»йҮҚе№¶дҝқз•ҷйЎәеәҸ
        return "_".join(values) if values else "жңӘе‘ҪеҗҚиЎҢ"

    def _to_dict(self):
        """
        е°Ҷ DataFrame иҪ¬жҚўдёәеӯ—е…ё {иЎҢеҗҚ_еҲ—еҗҚ: еҖј}
        """
        data_dict = {}
        for idx, row in self.df.iterrows():
            row_name = row["иЎҢеҗҚ"]
            for col in self.df.columns:
                if col in self.row_name_cols or col == "иЎҢеҗҚ" or col == "":
                    continue
                key = f"{row_name}_{col}"
                data_dict[key] = row[col]
        return data_dict

    def get_dict(self):
        """
        иҝ”еӣһ (иЎҢеҗҚ_еҲ—еҗҚ -> еҖј) зҡ„еӯ—е…ё
        """
        return self.data_dict

    def update_from_dict(self, updates: dict):
        """
        ж №жҚ® updates = {иЎҢеҗҚ_еҲ—еҗҚ: еҖј} жӣҙж–° DataFrame
        """
        for key, value in updates.items():
            *row_parts, col = key.split("_")
            row_name = "_".join(row_parts)

            # жүҫиЎҢ
            row_idx = self.df.index[self.df["иЎҢеҗҚ"] == row_name].tolist()
            if not row_idx:
                print(f"жңӘжүҫеҲ°иЎҢ: {row_name}")
                continue

            # жӣҙж–°еҖј
            self.df.at[row_idx[0], col] = value
            self.data_dict[key] = value  # еҗҢжӯҘжӣҙж–°еӯ—е…ё
            
    def get_df(self, drop_row_name=False):
        """
        иҝ”еӣһеҪ“еүҚ DataFrame
        :param drop_row_name: жҳҜеҗҰеҲ йҷӨеҶ…йғЁз”ҹжҲҗзҡ„ 'иЎҢеҗҚ' еҲ—
        """
        df_copy = self.df.copy()
        if drop_row_name and "иЎҢеҗҚ" in df_copy.columns:
            df_copy = df_copy.drop(columns=["иЎҢеҗҚ"])
        return df_copy
    
def excel_to_csv_cleaned(excel_file, mysqlwriter, chatbot, history):
    output_dir = "output_csv"
    os.makedirs("output_csv", exist_ok=True)
    # жү“ејҖ Excel ж–Үд»¶
    wb = load_workbook(excel_file, data_only=True)
    target_mapper, meta, header_row = None, None, None
    for sheet_name in wb.sheetnames:
        print(f"жӯЈеңЁеӨ„зҗҶ sheet: {sheet_name} ...")
        ws = wb[sheet_name]

        data = ws.values
        data1 = []
        for row in ws.values:
            clean_row = ["" if cell is None else cell for cell in row]
            data1.append(clean_row)
        df_org = pd.DataFrame(data1)
        df_org = df_org.dropna(how="all")
        for merged in ws.merged_cells.ranges:
            min_row, max_row = merged.min_row - 1, merged.max_row - 1
            min_col, max_col = merged.min_col - 1, merged.max_col - 1
            value = df_org.iat[min_row, min_col]
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    if df_org.iat[r, c] is None or str(df_org.iat[r, c]).strip() == "":
                        df_org.iat[r, c] = value

        df, header_row = detect_headers(df=df_org, keyword="еәҸеҸ·", max_header_rows=5)

        if df is None:
            print("и·іиҝҮеҪ“еүҚ sheetпјҲжңӘжЈҖжөӢеҲ°иЎЁеӨҙпјү")
            if "жұҮжҖ»иЎЁ" in sheet_name:
                target_mapper, meta, header_row = no_header_sheet(df_org, sheet_name)
            continue
        meta_info = extract_meta_info(df_org.iloc[:header_row])
        # еҲ йҷӨе®Ңе…Ёз©әиЎҢ
        df = df.dropna(how="all")
        first_col = df.columns[0]
        df = df[~df[first_col].astype(str).str.contains("еҗҲи®Ў|е°Ҹи®Ў", na=False)]
        df.columns = [
            str(c).strip() if c is not None and str(c).strip() != "" else f"жңӘе‘ҪеҗҚеҲ—_{i}"
            for i, c in enumerate(df.columns)
        ]
        df = df[[col for col in df.columns if "жңӘе‘ҪеҗҚеҲ—" not in col]]
        new_columns = []
        name_count = {}
        for col in df.columns:
            if col not in name_count:
                name_count[col] = 0
                new_columns.append(col)
            else:
                name_count[col] += 1
                new_columns.append(f"{col}пјҲ{name_count[col]}пјү")  # дҪҝз”Ёдёӯж–ҮжӢ¬еҸ·
        df.columns = new_columns
        text_cols, num_cols = split_text_num_cols(df)
        for col in num_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce').round(2)

        # еәҸеҸ·йҖ’еўһ
        if "еәҸеҸ·" in str(first_col):
            df[first_col] = range(1, len(df) + 1)
        else:
            df.insert(0, "еәҸеҸ·", range(1, len(df) + 1))

        df = replace_ref_with_nan(df)
        df = df.dropna(how="all", subset=[col for col in df.columns if col != "еәҸеҸ·"])
        # output_file = os.path.join(output_dir, f"{sheet_name}.csv")
        # df.to_csv(output_file, index=False, encoding="utf-8-sig")
        # print(f"е·Ідҝқеӯҳ: {output_file}")
        company = meta_info.get("company", "жңӘзҹҘе…¬еҸё")  # еҰӮжһңжІЎжңүпјҢе°ұз”Ёй»ҳи®ӨеҖј
        sheet_name = f"{company}_{sheet_name}"
        # check
        res = re.findall(r"(еҹәзЎҖиЎЁ\d+)", sheet_name)
        res = res[0]
        tables = mysqlwriter.find_tables_like(res)
        def df_to_compare_dict(df, key_col, value_cols):
            result = {}
            cols_to_use = [key_col] + value_cols
            for idx, row in df[cols_to_use].reset_index().iterrows():
                row_id = int(row["index"])      # иЎҢеҸ·
                key = row[key_col]
                
                if pd.isna(key) or str(key).strip() == "":
                    continue

                result[row_id] = {
                    "key": key,
                    "value": {col: row[col] for col in value_cols},
                    "row_index": row_id,
                    "history_compare": {}   # еҗҺз»ӯеЎ«е……еҺҶеҸІеҜ№жҜ”з»“жһң
                }
            return result
        
        compare_config = {
            "еҹәзЎҖиЎЁ3":{
                "key_col": "е…·дҪ“еІ—дҪҚ",
                "value_col": "е°Ҹи®Ў"
            },
            "еҹәзЎҖиЎЁ4":{
                "key_col": "и®ҫеӨҮеҗҚз§°",
                "value_col": "иҙӯзҪ®еҚ•д»·"
            },
            "еҹәзЎҖиЎЁ5":{
                "key_col": "йўҶз”Ёжқҗж–ҷеҗҚз§°",
                "value_col": "йўҶз”Ёжқҗж–ҷеҚ•д»·"
            },
            "еҹәзЎҖиЎЁ6":{
                "key_col": "иғҪиҖ—жҳҺз»Ҷ",
                "value_col": "иғҪиҖ—еҚ•д»·"
            },
            "еҹәзЎҖиЎЁ7":{
                "key_col": "е…¶д»–ж”ҜеҮәеҗҚз§°е’Ңз®Җиҝ°",
                "value_col": "иҙ№з”ЁйҮ‘йўқ"
            }
        }
        if res in compare_config:
            key_col = compare_config[res]['key_col']
            base_value_col = compare_config[res]['value_col']
            value_cols = [col for col in df.columns if base_value_col in col]
            if key_col not in df.columns or not value_cols:
                print(f"иЎЁ {table} дёҚеӯҳеңЁ {key_col} жҲ– {base_value_col} еӯ—ж®өпјҢи·іиҝҮ")
            else:
                result = df_to_compare_dict(df, key_col, value_cols)
                print(key_col, value_cols)
                for table in tables:
                    if "дёҠжө·" in table:
                        continue
                    df_hist = mysqlwriter.select_df(table)  # е…Ҳе…ЁиЎЁиҜ»еҸ–

                    hist_value_cols = [col for col in df_hist.columns if base_value_col in col]
                    if key_col not in df_hist.columns or not hist_value_cols:
                        print(f"иЎЁ {table} дёҚеӯҳеңЁ {key_col} жҲ– {base_value_col} еӯ—ж®өпјҢи·іиҝҮ")
                        continue

                    # еҶҚйҖүйңҖиҰҒзҡ„еҲ—
                    cols_to_use = [key_col] + hist_value_cols
                    df_hist = df_hist[cols_to_use].copy()
                    df_hist[key_col] = df_hist[key_col].astype(str)
                    hist_map = {col: defaultdict(list) for col in hist_value_cols}
                    for k, *vals in zip(df_hist[key_col], *[df_hist[col] for col in hist_value_cols]):
                        for col, v in zip(hist_value_cols, vals):
                            hist_map[col][str(k)].append(v)

                    # еЎ«е……еҺҶеҸІжҜ”еҜ№з»“жһң
                    for row_id, item in result.items():
                        key_value = str(item["key"])
                        for col in hist_value_cols:
                            item["history_compare"].setdefault(col, [])
                            if key_value in hist_map[col]:
                                item["history_compare"][col].append(hist_map[col][key_value])

                # print(json.dumps(result, ensure_ascii=False, indent=4))
                out_of_range_ids = {col: [] for col in value_cols}
                for row_id, item in result.items():
                    for col in value_cols:
                        current_value = safe_number(item["value"].get(col))
                        if current_value is None or pd.isna(current_value) or current_value == 0:
                            continue  # и·іиҝҮз©әеҖје’Ң0
                        all_history_values = []
                        if "history_compare" in item:
                            for hist_col, hist_lists in item["history_compare"].items():
                                for sublist in hist_lists:
                                    for v in sublist:
                                        v_num = safe_number(v)
                                        if v_num is not None:
                                            all_history_values.append(v_num)

                        if not all_history_values:
                            continue  # еҰӮжһңеҺҶеҸІеҖје…ЁжҳҜз©әжҲ–0пјҢи·іиҝҮ

                        # и®Ўз®—е№іеқҮеҖј
                        avg_history = sum(all_history_values) / len(all_history_values)

                        # еҲӨж–ӯжҳҜеҗҰи¶…еҮәиҢғеӣҙ
                        if not (avg_history * 0.8 <= current_value <= avg_history * 1.2):
                            out_of_range_ids[col].append(row_id + header_row)

                # жүҫеҲ° value_col еҜ№еә”зҡ„еҲ—зҙўеј•пјҲDataFrameеҲ—еҗҚ -> wsеҲ—зҙўеј•пјү
                fill_red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

                # out_of_range_ids жҳҜеҹәдәҺ DataFrame зҡ„иЎҢеҸ·пјҢйңҖиҰҒеҮҸеҺ» header_row еҶҚеҠ дёҠе®һйҷ… Excel иө·е§ӢиЎҢ
                for col in value_cols:
                    value_col_idx = df.columns.get_loc(col)  # еҲ—зҙўеј•
                    for row_id in out_of_range_ids[col]:
                        excel_row = row_id + header_row
                        ws.cell(row=excel_row, column=value_col_idx + 1).fill = fill_red

    output_excel_file = os.path.join("output_excel", os.path.basename(excel_file).replace(".xlsx", "_marked.xlsx"))
    os.makedirs("output_excel", exist_ok=True)
    chatbot.append(["ж–Үд»¶дҝқеӯҳжҲҗеҠҹ", output_excel_file])
    wb.save(output_excel_file)
    promote_file_to_downloadzone(output_excel_file, chatbot=chatbot)
    chatbot.append(["е·Із”ҹжҲҗдёӢиҪҪж–Үд»¶", f"и¶…еҮәеҺҶеҸІиҢғеӣҙзҡ„еҚ•е…ғж је·Іж Үзәў"])
    yield from update_ui(chatbot=chatbot, history=history)
    return target_mapper, meta, header_row

def safe_number(x):
    """
    е°ҶеҖјиҪ¬жҚўдёә floatпјҢе№¶и·іиҝҮ None, з©әеӯ—з¬ҰдёІ, 0
    """
    if x is None:
        return None
    try:
        num = float(str(x).replace(",", "").strip())
    except ValueError:
        return None
    if num == 0 or math.isnan(num):
        return None
    return num

def split_text_num_cols(df: pd.DataFrame):
    text_cols, num_cols = [], []
    for col in df.columns:
        series = df[col].dropna().astype(str).str.strip()
        # иҝҮж»ӨжҺүз©әеҖјгҖҒж–ңжқ жҲ–зү№ж®ҠеҚ дҪҚз¬Ұ
        series = series[~series.isin(['', '/', '-'])]
        if len(series) == 0:
            text_cols.append(col)
            continue
        # еҲӨж–ӯеү©дҪҷеҖјжҳҜеҗҰе…Ёдёәж•°еҖј
        is_num = series.str.match(r'^[+-]?(\d+(\.\d+)?([eE][+-]?\d+)?)$').all()
        if is_num:
            num_cols.append(col)
        else:
            text_cols.append(col)
    print(f"ж–Үжң¬еҲ—: {text_cols}")
    print(f"ж•°еҖјеҲ—: {num_cols}")
    return text_cols, num_cols

def replace_ref_with_nan(df: pd.DataFrame) -> pd.DataFrame:
    """
    жЈҖжҹҘ DataFrame дёӯжҳҜеҗҰжңү '#REF!'пјҢеҰӮжһңжңүеҲҷжӣҝжҚўдёә np.nanгҖӮ
    """
    def clean_cell(x):
        if x == '#REF!':
            return np.nan
        return x

    return df.map(clean_cell)

def detect_headers(df, keyword="еәҸеҸ·", max_header_rows=5):
    """
    иҮӘеҠЁжЈҖжөӢе№¶еҗҲе№¶еӨҡиЎҢиЎЁеӨҙ
    еҸӮж•°:
        df: DataFrame
        keyword: иЎЁеӨҙиҜҶеҲ«е…ій”®иҜҚ (й»ҳи®Өжүҫ "еәҸеҸ·")
        max_header_rows: жңҖеӨ§иЎЁеӨҙиЎҢж•° (й»ҳи®Ө5)
    иҝ”еӣһ:
        (еӨ„зҗҶеҗҺзҡ„df, иЎЁеӨҙиЎҢж•°)
    """
    header_row = None
    # е…ҲжүҫеҲ°еҗ«жңүе…ій”®еӯ—зҡ„йӮЈдёҖиЎҢ
    for i in range(min(10, len(df))):
        if df.iloc[i].astype(str).str.contains(keyword).any():
            header_row = i
            break

    if header_row is None:
        return None, None

    # еҲқе§ӢиЎЁеӨҙ
    headers = [df.iloc[header_row].astype(str).tolist()]
    next_row = header_row + 1

    while next_row < len(df) and len(headers) < max_header_rows:
        row = df.iloc[next_row].astype(str)
        first_cell = row.iloc[0].strip()

        if first_cell.isdigit():
            break
        elif first_cell in ["еҗҲи®Ў", "е°Ҹи®Ў", "жҖ»и®Ў"]:
            next_row += 1
            continue
        else:
            headers.append(row.tolist())
            next_row += 1

    # еҗҲе№¶еӨҡиЎҢиЎЁеӨҙ
    final_header = []
    for col_values in zip(*headers):
        col_values = [v for v in col_values if v not in ["nan", "None", ""]]
        cleaned = []
        for v in col_values:
            if not cleaned or cleaned[-1] != v:
                cleaned.append(v)
        final_header.append("_".join(cleaned) if len(cleaned) > 1 else (cleaned[0] if cleaned else ""))

    # и®ҫзҪ®еҲ—еҗҚ
    df.columns = final_header

    # еҰӮжһңжҳҜ MultiIndexпјҢд№ҹиҪ¬жҚўжҲҗеӯ—з¬ҰдёІ
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = ["_".join([str(c) for c in col if c not in ["nan","None",""]]) 
                      for col in df.columns]
    else:
        df.columns = [str(c) for c in df.columns]

    # еҺ»жҺүиЎЁеӨҙиЎҢ
    df = df.iloc[header_row + len(headers):].reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]

    # еҲ йҷӨвҖңеҲ—еҗҚдёәз©әдё”иҜҘеҲ—е…Ёдёәз©әвҖқзҡ„еҲ—
    cols_to_keep = []
    for col in df.columns:
        # еҰӮжһң df[col] иҝ”еӣһ DataFrameпјҲйҮҚеӨҚеҲ—еҗҚжғ…еҶөпјүпјҢеҸ–з¬¬дёҖеҲ—
        series = df[col]
        if isinstance(series, pd.DataFrame):
            series = series.iloc[:, 0]

        # дҝқз•ҷи§„еҲҷпјҡеҲ—еҗҚйқһз©әпјҢжҲ–иҖ…еҲ—йҮҢжңүж•°жҚ®
        if col not in ['', 'None'] or not series.isna().all():
            cols_to_keep.append(col)

    df = df[cols_to_keep].copy()
    return df, header_row

def extract_meta_info(df):
    """
    д»ҺиЎЁеӨҙеүҚзҡ„иЎҢжҸҗеҸ–е…¬еҸёеҗҚз§°е’Ңcomment
    """
    meta = {"company": None, "comment": None}
    rows = df.astype(str).fillna("").iloc[:, 1].tolist()  

    for row in rows:
        row = row.strip()
        if not row:
            continue

        # жҸҗеҸ–е…¬еҸёеҗҚз§°
        if "е…¬еҸёеҗҚз§°" in row:
            meta["company"] = row.split("пјҡ", 1)[-1].strip()

        # жҸҗеҸ– comment
        if "еЎ«иЎЁиҜҙжҳҺ" in row:
            meta["comment"] = row
    
    return meta

def no_header_sheet(df_org, sheet_name):
    output_flag = "ж”ҜеҮәжұҮжҖ»иЎЁ" in sheet_name

    header_row = None
    for i in range(len(df_org)):
        first_col_value = str(df_org.iloc[i, 0]).strip()
        if "ж”ҜеҮәзұ»еһӢ" in first_col_value:
            header_row = i
            break
    if header_row is None:
        print(f"жңӘжүҫеҲ°иЎЁеӨҙпјҡ{sheet_name}")
        return None

    meta = extract_meta_info(df_org.iloc[:header_row])
    
    df = df_org.iloc[header_row:].reset_index(drop=True)
    df = df.dropna(axis=1, how="all")
    mapper = SheetMapper(df)
    company = meta.get("company", "жңӘзҹҘе…¬еҸё")  # еҰӮжһңжІЎжңүпјҢе°ұз”Ёй»ҳи®ӨеҖј
    sheet_name = f"{company}_{sheet_name}"
    if output_flag:
        return mapper, meta, header_row
    else:
        return None, None, None # иҝ”еӣһеӨ„зҗҶеҗҺзҡ„ DataFrame
    
@CatchException
def еҺҶеҸІж•°жҚ®еҜ№жҜ”(main_input, llm_kwargs, plugin_kwargs, chatbot, history, system_prompt, user_request):
    chatbot.append(["еҗҜеҠЁж•°жҚ®дёҠдј д»»еҠЎ", "иҜ»еҸ–ж–Үд»¶е№¶иҮӘеҠЁеӨ„зҗҶsheetдёӯ..."])
    yield from update_ui(chatbot=chatbot, history=history)
    if os.path.isdir(main_input):
        # еҸ–зӣ®еҪ•дёӢз¬¬дёҖдёӘ Excel ж–Үд»¶
        files = glob.glob(os.path.join(main_input, "*.xlsx"))
        if not files:
            chatbot.append(["й”ҷиҜҜ", "зӣ®еҪ•дёӯжІЎжңүжүҫеҲ°д»»дҪ• .xlsx ж–Үд»¶"])
            yield from update_ui(chatbot=chatbot, history=history)
            return
        excel_file = files[0]  # еҸӘеҸ–з¬¬дёҖдёӘж–Үд»¶
    else:
        if not os.path.isfile(main_input):
            chatbot.append(["й”ҷиҜҜ", f"ж–Үд»¶дёҚеӯҳеңЁпјҡ{main_input}"])
            yield from update_ui(chatbot=chatbot, history=history)
            return
        excel_file = main_input  # main_input жң¬иә«е°ұжҳҜж–Үд»¶
    
    chatbot.append(["иҜ»еҸ–ж–Үд»¶", os.path.basename(excel_file)])
    yield from update_ui(chatbot=chatbot, history=history)

    total_dict = {}
    mysqlwriter = MySQLWriter("lifeng")
    target_mapper, meta, header_row = yield from excel_to_csv_cleaned(excel_file, mysqlwriter, chatbot, history)

